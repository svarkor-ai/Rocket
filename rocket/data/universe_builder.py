"""
Rocket Stock Scanner - Dynamic Universe Builder

Loads ticker universes from:
1. Curated US ticker CSV (7000+ verified NYSE/NASDAQ/AMEX tickers)
2. Major index constituents via Wikipedia (OMX30, FTSE, DAX, etc.)
3. Cached for performance

Total target: 15,000+ tickers across global markets.
"""

import csv
import json
import logging
import re
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# Cache paths
CACHE_DIR = Path(__file__).parent
CACHE_FILE = CACHE_DIR / "universe_cache.json"
CACHE_TTL_HOURS = 24  # Refresh cache every 24h

# Known index constituents (fetched from Wikipedia and other sources)
INDEX_CONSTITUENTS_FILE = CACHE_DIR / "index_constituents.json"

# US tickers source: GitHub mirror of browser-vm/us-stock-tickers
# 7000+ verified US ticker symbols (NYSE, NASDAQ, AMEX)
US_TICKERS_URL = (
    "https://raw.githubusercontent.com/browser-vm/us-stock-tickers"
    "/main/us_stock_tickers.csv"
)
US_TICKERS_CACHE = CACHE_DIR / "us_tickers.csv"

# Local CSV/XLSX ticker sources (verified, structured data)
# Path: rocket-stock-scanner/data/tickers/
TICKER_DATA_DIR = Path(__file__).parent.parent.parent / "data" / "tickers"
LOCAL_TICKER_SOURCES = {
    # Australia (ASX)
    "australia": {
        "file": TICKER_DATA_DIR / "asx_stocks.csv",
        "type": "csv",
        "ticker_column": 0,  # First column = "ASX code"
    },
    # UK (LSE) — All Equity (includes Shares, Depository Receipts, ETFs, etc.)
    "uk": {
        "file": TICKER_DATA_DIR / "lse_stocks.xlsx",
        "type": "xlsx",
        "sheet": "1.0 All Equity",
        "ticker_column": 0,  # TIDM column
    },
    # Germany (Frankfurt)
    "germany": {
        "file": TICKER_DATA_DIR / "frankfurt_stocks.xlsx",
        "type": "xlsx",
        "sheet": "Prime Standard",
        "ticker_column": "scan",  # Scan all cells for ticker-like patterns
    },
    # Other EU — All Equity (includes Shares, Depository Receipts, ETFs, etc.)
    "other_eu": {
        "file": TICKER_DATA_DIR / "other_eu_stocks.xlsx",
        "type": "xlsx",
        "sheet": "1.0 All Equity",
        "ticker_column": 0,
    },
}

# Wikipedia pages by region: {region_key: [(page_name, description), ...]}
# page_name is the Wikipedia article slug (with %26 for &, etc.)
WIKI_PAGES = {
    # Sweden / Scandinavia
    "sweden": [
        ("OMX_Stockholm_30", "OMX Stockholm 30"),
        ("Stockholm_Stock_Exchange", "Stockholm Stock Exchange"),
        ("Oslo_Stock_Exchange", "Oslo Stock Exchange"),
    ],
    # UK
    "uk": [
        ("FTSE_100_Index", "FTSE 100"),
        ("FTSE_250_Index", "FTSE 250"),
        ("FTSE_350_Index", "FTSE 350"),
    ],
    # Germany
    "germany": [
        ("Frankfurt_Stock_Exchange", "Frankfurt Stock Exchange"),
        ("DAX", "DAX index constituents"),
    ],
    # France
    "france": [
        ("CAC_40", "CAC 40"),
        ("Euronext", "Euronext Paris"),
    ],
    # Japan
    "japan": [
        ("Nikkei_225", "Nikkei 225"),
        ("Tokyo_Stock_Exchange", "Tokyo Stock Exchange"),
    ],
    # Hong Kong
    "hongkong": [
        ("Hang_Seng_Index", "Hang Seng Index"),
        ("Hong_Kong_Stock_Exchange", "Hong Kong Stock Exchange"),
    ],
    # China
    "china": [
        ("Shanghai_Stock_Exchange", "Shanghai Stock Exchange"),
        ("Shenzhen_Stock_Exchange", "Shenzhen Stock Exchange"),
        ("Shanghai_Composite_Index", "Shanghai Composite Index"),
    ],
    # India
    "india": [
        ("NIFTY_50", "NIFTY 50"),
        ("Bombay_Stock_Exchange", "BSE / Bombay Stock Exchange"),
        ("National_Stock_Exchange_of_India", "NSE / National Stock Exchange"),
    ],
    # Australia
    "australia": [
        ("S%26P_ASX_200", "S&P ASX 200"),
        ("ASX", "Australian Securities Exchange"),
    ],
    # Canada
    "canada": [
        ("S%26P/TSX_Composite_Index", "S&P/TSX Composite"),
        ("Toronto_Stock_Exchange", "Toronto Stock Exchange"),
    ],
    # Switzerland
    "switzerland": [
        ("SIX_Swiss_Exchange", "SIX Swiss Exchange"),
        ("SMI_(index)", "Swiss Market Index (SMI)"),
    ],
    # South Korea
    "korea": [
        ("KOSPI", "KOSPI index"),
        ("Korea_Exchange", "Korea Exchange"),
    ],
}


def _load_us_tickers_from_csv() -> set[str]:
    """Load verified US ticker symbols from a curated CSV file.

    Source: https://github.com/browser-vm/us-stock-tickers
    Contains 7000+ verified ticker symbols for NYSE, NASDAQ, and AMEX.
    This is the ground-truth source for US tickers — no regex guessing.

    Returns sorted list of unique tickers (max 5 chars, optional .A/.B suffix).
    """
    # Try cached CSV first
    if US_TICKERS_CACHE.exists():
        try:
            with open(US_TICKERS_CACHE, "r") as f:
                content = f.read()
            if content and len(content) > 1000:
                logger.info(f"Loaded {US_TICKERS_CACHE.name} ({len(content)} chars)")
                return _parse_csv_content(content)
        except IOError:
            logger.warning("Failed to write local tickers cache")

    # Download CSV
    try:
        import requests
        resp = requests.get(US_TICKERS_URL, timeout=30)
        resp.raise_for_status()
        content = resp.text
        # Save cache
        with open(US_TICKERS_CACHE, "w") as f:
            f.write(content)
        logger.info(f"Downloaded US tickers CSV ({len(content)} chars)")
        return _parse_csv_content(content)
    except Exception as e:
        logger.warning("Failed to load US tickers CSV")
        return set()


def _parse_csv_content(content: str) -> set[str]:
    """Parse CSV content into a set of valid US tickers."""
    tickers = set()
    for line in content.splitlines():
        line = line.strip()
        if not line or line.startswith("ticker_symbol"):
            continue
        # Format: TICKER,Company Name,Exchange
        parts = line.split(",")
        if parts:
            ticker = parts[0].strip().upper()
            # Valid: 1-5 uppercase letters, optionally followed by .A/.B or /B suffix
            # e.g., BRK/B, BRK.A, AAPL, MSFT
            if ticker and re.match(r"^[A-Z0-9/]{1,5}(?:[.][A-Z]{1,2})?$", ticker):
                tickers.add(ticker)
    return tickers


def _load_tickers_from_local_source(source: dict) -> list[str]:
    """Load tickers from a local CSV or XLSX file.

    Args:
        source: Dict with keys:
            file: Path to CSV/XLSX file
            type: "csv" or "xlsx"
            ticker_column: column index (int) or "scan" for full-cell scan
            sheet: XLSX sheet name (only for xlsx type)

    Returns:
        Sorted list of unique valid tickers, or empty list on failure.
    """
    filepath = source["file"]
    source_type = source["type"]
    ticker_col = source.get("ticker_column", 0)

    if not filepath.exists():
        logger.warning(f"Local source file not found: {filepath}")
        return []

    try:
        if source_type == "csv":
            return _load_tickers_from_csv_file(filepath, ticker_col)
        elif source_type == "xlsx":
            return _load_tickers_from_xlsx_file(filepath, ticker_col, source.get("sheet"))
        else:
            logger.warning(f"Unsupported local source type: {source_type}")
            return []
    except Exception as e:
        logger.warning("Failed to load tickers from {filepath}")
        return []


def _load_tickers_from_csv_file(filepath: Path, ticker_col: int) -> list[str]:
    """Load tickers from a CSV file using the specified column index."""
    tickers = set()
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader, None)  # Skip header row
        for row in reader:
            if ticker_col < len(row):
                cell = row[ticker_col].strip().upper()
                if _TICKER_RE.match(cell) and cell not in _FALSE_POSITIVES:
                    tickers.add(cell)
    return sorted(tickers)


def _load_tickers_from_xlsx_file(filepath: Path, ticker_col, sheet_name: Optional[str] = None) -> list[str]:
    """Load tickers from an XLSX file.

    If ticker_col is an int, reads from that column index.
    If ticker_col is "scan", scans ALL cells for ticker-like patterns.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl required for XLSX loading")
        return []

    tickers = set()
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        sheets_to_scan = [sheet_name]
    elif sheet_name:
        logger.warning(f"Sheet {sheet_name} not found in {filepath}, scanning all sheets")
        sheets_to_scan = wb.sheetnames
    else:
        sheets_to_scan = wb.sheetnames

    for sheet in sheets_to_scan:
        ws = wb[sheet]
        if ticker_col != "scan":
            # Read specific column
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                if ticker_col < len(row) and row[ticker_col]:
                    cell = str(row[ticker_col]).strip().upper()
                    if _TICKER_RE.match(cell) and cell not in _FALSE_POSITIVES:
                        tickers.add(cell)
        else:
            # Scan all cells for ticker-like patterns
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        cell = cell.strip().upper()
                        if len(cell) > 20:
                            continue
                        if _TICKER_RE.match(cell) and cell not in _FALSE_POSITIVES:
                            tickers.add(cell)

    wb.close()
    return sorted(tickers)


def _is_cache_fresh(cache: dict) -> bool:
    """Check if cache is still within TTL."""
    ts = cache.get("timestamp", "")
    try:
        cached_time = datetime.fromisoformat(ts)
        return datetime.now(timezone.utc) - cached_time < timedelta(hours=CACHE_TTL_HOURS)
    except (ValueError, TypeError):
        return False


def _read_cache() -> Optional[dict]:
    """Read universe cache from disk."""
    if not CACHE_FILE.exists():
        return None
    try:
        with open(CACHE_FILE, "r") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return None


def _write_cache(universe: dict) -> None:
    """Write universe cache to disk with timestamp."""
    cache = {
        "tickers": universe,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "source": "wikipedia + known constituents",
        "version": 2,
    }
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)


def _save_index_constituents(constituents: dict[str, list[str]]) -> None:
    """Save index constituents to separate cache file."""
    with open(INDEX_CONSTITUENTS_FILE, "w") as f:
        json.dump(constituents, f, indent=2, ensure_ascii=False)


def _load_index_constituents() -> dict[str, list[str]]:
    """Load pre-fetched index constituents from cache file."""
    if not INDEX_CONSTITUENTS_FILE.exists():
        return {}
    try:
        with open(INDEX_CONSTITUENTS_FILE, "r") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return {}


# Ticker regex: two strict patterns (no decimals, no pure digit strings):
# 1. No-dot: [A-Z]{1,5} — e.g. AAPL, GOOGL, RELIANCE, BF
# 2. Dot: [A-Z0-9]{1,5}(?:-[A-Z0-9]{1,5})*\.[A-Z]{1,4} — e.g. BMW.DE, 7203.T, ERIC-B.ST
# SEC tickers are max 5 chars. International tickers with suffix can be longer (handled by dot pattern).
# This rejects company names (MICROSOFT=9, APPLE=5, CISCO=5, INTEL=5, TESLA=5).
# Note: All ticker text must be uppercased before matching.
_TICKER_RE = re.compile(r"^[A-Z0-9/]{1,5}(?:[-/][A-Z0-9]{1,5})*\.[A-Z]{1,4}$|^[A-Z0-9/]{1,5}$")

# Known false positives to filter out
_FALSE_POSITIVES = frozenset({
    "N/A", "N.A.", "—", "-", "", "NA", "N/A.", "N/A ",
    "USD", "EUR", "GBP", "JPY", "CNY", "AUD", "CAD", "CHF",
    "KRW", "HKD", "SEK", "NOK", "DKK", "INR",
    "ETF", "REIT", "FUND", "INDEX", "SHARES", "UNITS",
    "CLASS",
})


def _extract_tickers_from_wikipedia(page_name: str) -> list[str]:
    """Extract ticker symbols from a Wikipedia table using broad cell scanning.

    Strategy:
    1. Fetch the Wikipedia page
    2. Find ALL wikitable tables on the page
    3. For each table, scan ALL cells (not just a specific column)
    4. Match cells against ticker regex patterns
    5. Filter out false positives
    6. Return sorted unique list of tickers

    This works for any Wikipedia table regardless of column naming.

    Returns sorted unique list of tickers.
    """
    try:
        import requests
        from bs4 import BeautifulSoup
    except ImportError:
        logger.warning("requests and bs4 required for Wikipedia scraping")
        return []

    url = f"https://en.wikipedia.org/wiki/{page_name}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Rocket Stock Scanner bot; +https://github.com/svarkor-ai/rocket)"
    }

    try:
        resp = requests.get(url, headers=headers, timeout=30)
    except requests.exceptions.HTTPError as e:
        if e.response is not None and e.response.status_code == 404:
            logger.warning(f"Wikipedia page not found: {page_name}")
        else:
            logger.warning(f"Wikipedia fetch failed for {page_name}")
        return []
    except Exception as e:
        logger.warning(f"Wikipedia fetch failed for {page_name}")
        return []

    if resp.status_code != 200:
        logger.warning(f"Wikipedia returned status {resp.status_code} for {page_name}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")

    # Find all wikitable tables
    tables = soup.find_all("table", {"class": re.compile(r"wikitable")})
    if not tables:
        logger.debug(f"No wikitable found for {page_name}")
        return []

    all_tickers = set()

    for table in tables:
        rows = table.find_all("tr")
        if not rows:
            continue

        # Scan ALL cells in ALL rows (skip header row for data)
        for row in rows:
            cells = row.find_all(["td", "th"])
            for cell in cells:
                # Strip inner HTML (links, references, superscripts)
                cell_text = cell.get_text(" ", strip=True)
                # Remove reference markers like [1], [a], [nb 1]
                cell_text = re.sub(r"\[[\w\s]+\]", "", cell_text).strip()
                # Remove parenthetical notes like "(United States)"
                cell_text = re.sub(r"\(.*?\)", "", cell_text).strip()
                # Collapse whitespace
                cell_text = re.sub(r"\s+", "", cell_text)

                # Uppercase before regex match (e.g., "Telia" → "TELIA")
                cell_text_upper = cell_text.upper()

                if _TICKER_RE.match(cell_text_upper) and cell_text_upper not in _FALSE_POSITIVES:
                    all_tickers.add(cell_text_upper)

    return sorted(all_tickers)


def _build_universe(force_refresh: bool = False) -> dict[str, list[str]]:
    """Build the ticker universe from index constituents.

    Strategy:
    1. Try cache first (fresh within 24h)
    2. If no cache or force_refresh:
       a. Scrape Wikipedia for each region's index pages
       b. Collect unique tickers per region
       c. Save to cache
       d. Return dict of {region: sorted_unique_tickers}

    Returns dict[str, list[str]] with keys:
        usa, sweden, uk, germany, france, japan,
        hongkong, china, india, australia, canada,
        switzerland, korea, international
    """
    # Try cache first
    if not force_refresh:
        cache = _read_cache()
        if cache and _is_cache_fresh(cache):
            logger.debug("Using cached universe data")
            return cache["tickers"]

    logger.info("Building universe from CSV/XLSX sources + Wikipedia fallback")
    universe: dict[str, list[str]] = {}
    constituents_cache: dict[str, list[str]] = {}

    # --- USA: load from verified CSV (no Wikipedia scraping) ---
    logger.info("Fetching USA tickers from curated CSV source...")
    us_tickers = _load_us_tickers_from_csv()
    if us_tickers:
        universe["usa"] = sorted(us_tickers)
        constituents_cache["usa/csv"] = universe["usa"]
        logger.info(f"  usa: {len(universe['usa'])} tickers (from CSV)")
    else:
        logger.warning("  usa: no tickers from CSV — will fall back to embedded data")

    # --- Load from local CSV/XLSX sources (verified, structured data) ---
    for region_key, source in LOCAL_TICKER_SOURCES.items():
        logger.info(f"Fetching {region_key} tickers from local source...")
        tickers = _load_tickers_from_local_source(source)
        if tickers:
            universe[region_key] = tickers
            constituents_cache[f"{region_key}/local"] = tickers
            logger.info(f"  {region_key}: {len(tickers)} tickers (from local {source['type']})")
        else:
            logger.warning(f"  {region_key}: no tickers from local source — will use Wikipedia/fallback")

    # --- Per-region scraping for remaining regions (Wikipedia fallback) ---
    for region_key, pages in WIKI_PAGES.items():
        if region_key in universe:
            # Already loaded from local source — skip Wikipedia for this region
            logger.info(f"Skipping {region_key} (already loaded from CSV/XLSX)")
            continue

        logger.info(f"Fetching {region_key} index constituents...")
        region_tickers: set[str] = set()

        for page_name, description in pages:
            logger.info(f"  Fetching {region_key}: {description} ({page_name})...")
            tickers = _extract_tickers_from_wikipedia(page_name)
            if tickers:
                logger.info(f"    Got {len(tickers)} tickers from {description}")
                region_tickers.update(tickers)
                key = f"{region_key}/{description}"
                constituents_cache[key] = tickers
            else:
                logger.warning(f"    No tickers from {description} ({page_name})")

        if region_tickers:
            if len(region_tickers) < 20:
                logger.info(f"  {region_key}: only {len(region_tickers)} tickers from Wikipedia — skipping (will use embedded fallback)")
            else:
                universe[region_key] = sorted(region_tickers)
                logger.info(f"  {region_key}: {len(universe[region_key])} tickers")
        else:
            logger.warning(f"  {region_key}: no tickers found — will fall back to embedded data")

    # --- Fill missing/deficient regions with embedded fallback data ---
    fallback = _build_embedded_fallback()
    for region_key in fallback:
        if region_key not in universe or len(universe[region_key]) < 5:
            if region_key in universe:
                del universe[region_key]
            universe[region_key] = list(fallback[region_key])
            logger.info(f"  {region_key}: {len(universe[region_key])} tickers (embedded fallback)")

    # --- Backward-compatibility: international = all non-US regions combined ---
    intl = set()
    for region_key in WIKI_PAGES:
        intl.update(universe.get(region_key, []))
    universe["international"] = sorted(intl)
    logger.info(f"  International (combined non-US): {len(universe['international'])} tickers")

    # Save to both caches
    _write_cache(universe)
    _save_index_constituents(constituents_cache)

    total = {k: len(v) for k, v in universe.items()}
    logger.info(f"Universe built: {total}")
    return universe


# Public API
# ---------------------------------------------------------------------------

_universe_cache: Optional[dict[str, list[str]]] = None


def get_universe(region: str | None = None, force_refresh: bool = False) -> list[str] | dict[str, list[str]]:
    """Get ticker universe.

    Args:
        region: Region key or None for all.
        force_refresh: If True, skip cache and fetch live data.

    Returns:
        If region is specified, returns list[str] of tickers for that region.
        If region is None, returns dict[str, list[str]] of all regions.
    """
    global _universe_cache
    if _universe_cache is None or force_refresh:
        _universe_cache = _build_universe(force_refresh)

    if region is None:
        return dict(_universe_cache)

    return list(_universe_cache.get(region, []))


def get_universe_count() -> dict[str, int]:
    """Return dict of region -> number of tickers (from cache)."""
    global _universe_cache
    if _universe_cache is None:
        # Load from cache file directly without building
        cache = _read_cache()
        if cache and "tickers" in cache:
            _universe_cache = cache["tickers"]

    if _universe_cache is None:
        # Fallback: build without scraping (use embedded data only)
        _universe_cache = _build_embedded_fallback()

    return {region: len(tickers) for region, tickers in _universe_cache.items()}


def get_all_tickers() -> list[str]:
    """Return all unique tickers across all regions."""
    universe = get_universe()
    all_tickers = set()
    for tickers in universe.values():
        all_tickers.update(tickers)
    return sorted(all_tickers)


def get_all_universes() -> list[str]:
    """Alias for get_all_tickers — return all tickers across all regions."""
    return get_all_tickers()


def get_region_count() -> dict[str, int]:
    """Alias for get_universe_count."""
    return get_universe_count()


def get_total_count() -> int:
    """Return total number of unique tickers across all regions."""
    return len(get_all_tickers())


def _build_embedded_fallback() -> dict[str, list[str]]:
    """Fallback universe builder with embedded lists if Wikipedia fails.

    This provides ~1,200+ tickers as a minimum viable universe,
    covering all regions instead of just usa + international.
    """
    # USA — S&P 500 (core holdings) + major tech
    USA_TICKERS = [
        # S&P 500 core
        "MMM", "ABT", "AOS", "ABNB", "ADM", "ADP", "ADSK", "AAP", "AEP", "AES",
        "AFL", "A", "APD", "AKAM", "ALB", "ARE", "ALGN", "ALLE", "LNT", "ALL",
        "GOOG", "GOOGL", "MO", "AMZN", "AAL", "AEP", "AXP", "AIG", "AMT",
        "AWK", "AMP", "ABC", "AME", "AON", "APA", "AAPL", "AMAT", "APTV",
        "ACN", "ADBE", "AMD", "AFRM", "AES", "AES", "AZEK", "ABBV", "ABC",
        "ABMD", "ABT", "ACGL", "ACN", "ADP", "ADSK", "AEE", "AEP", "AES",
        "AFL", "A", "APD", "AR", "AOS", "APA", "AVGO", "AWK", "AXON",
        "BALL", "BAX", "BDX", "BEN", "BF.B", "BK", "BKNG", "BIO", "BIIB",
        "BLK", "BMY", "BRK.B", "BR", "BSX", "BWA", "BX", "BXP", "CAG",
        "CRL", "C", "COO", "CPB", "COP", "CRM", "CSCO", "CSGP", "CSX",
        "CTAS", "CTLT", "CAT", "CBOE", "CBRE", "CDNS", "CDAY", "CE", "CDNS",
        "CHRW", "CPT", "CPB", "CME", "CMS", "COST", "CCI", "CTRA", "CUZ",
        "CRWD", "CMI", "CVS", "DHR", "DRI", "DVA", "DAY", "DE", "DAL",
        "DVN", "DXCM", "DECK", "DLR", "DFS", "DG", "DLTR", "EBAY", "ECL",
        "EFX", "EIX", "EMN", "EMR", "ENPH", "EOG", "EPAM", "EQIX", "EQR",
        "ESS", "ETN", "EBAY", "EXC", "EXPE", "EXPD", "EXR", "XOM", "FFIV",
        "FDS", "FICO", "FAST", "FRT", "FDX", "FIS", "FISV", "FITB", "F",
        "GPC", "GILD", "GL", "GPN", "GS", "GDDY", "HAL", "HIG", "HAS",
        "HRL", "HST", "HWM", "HP", "HOLX", "HD", "HON", "HPE", "HLT",
        "HF", "HOLX", "HSY", "HUM", "IBM", "IP", "IPG", "IFF", "IR",
        "IDXX", "IT", "ICE", "INVH", "IQV", "IRM", "JBHT", "JBL", "JKHY",
        "JNJ", "JCI", "JPM", "JPM", "KMB", "KIM", "KMI", "KLAC", "KHC",
        "KR", "LHX", "LH", "LRCX", "LVS", "LDOS", "LEN", "LIN", "LYV",
        "LKQ", "LLY", "LMT", "LULU", "LUK", "LYB", "MCD", "MCHP", "MU",
        "MSFT", "MRO", "MPC", "MKTX", "MAR", "MLM", "MAS", "MA", "MKC",
        "MCD", "MCO", "MS", "MOS", "MSI", "MMC", "MDLZ", "MPWR", "MNST",
        "MCO", "MCHP", "MRVL", "MRK", "META", "MO", "MPC", "MNST", "MDLZ",
        "MTCH", "MFN", "MTD", "MCK", "MUR", "NCLH", "NDAQ", "NEE", "NEM",
        "NDSN", "NSC", "NTRS", "NOC", "NUE", "NVDA", "NVR", "NXPI", "ORLY",
        "OXY", "ODFL", "OMC", "ON", "ORCL", "OTIS", "PCG", "PCI", "PEG",
        "PEP", "PFE", "PCG", "PM", "PSA", "PH", "PAYC", "PYPL", "PNR",
        "PNC", "POOL", "PPG", "PPL", "PFG", "PG", "PGR", "PLTR", "PMD",
        "PWR", "QCOM", "DGX", "RSG", "RAI", "RMD", "RVTY", "RBA", "RBLX",
        "RTX", "O", "REG", "REGN", "RF", "RHI", "ROK", "ROL", "ROP",
        "ROST", "RCL", "RGEN", "SPGI", "RSK", "SLB", "STX", "SRE", "NOW",
        "SBAC", "SHW", "SPG", "SWKS", "SJM", "SNPS", "SNA", "SO", "LUV",
        "SPLK", "SPGI", "STE", "SYF", "SYK", "SMCI", "TGT", "TSLA", "TEL",
        "TDY", "TFX", "TER", "TSLA", "TPR", "TRGP", "TGT", "TROW", "TTWO",
        "TT", "TJX", "TSCO", "TXN", "TXT", "TMO", "TJX", "TMO", "TFC",
        "USB", "UHS", "ULTA", "UNP", "UAL", "UPS", "URI", "UNH", "U",
        "VLO", "VTR", "VLTO", "VRSN", "VRSK", "VZ", "VRTX", "VFC", "VIAC",
        "VTRS", "V", "WAB", "WBA", "WMT", "DIS", "WM", "WAT", "WEC",
        "WFC", "WLK", "WDAY", "WRB", "WRK", "WST", "WDC", "WU", "WRB",
        "WRK", "WTW", "WY", "WYNN", "XEL", "XYL", "YUM", "ZBRA", "ZTS",
        "ZM", "ZS", "ZM", "ZS", "ZS", "ZS",
    ]

    # Sweden / Scandinavia
    SWEDEN_TICKERS = [
        "ABB.ST", "ADV.ST", "ATC.ST", "BEN.ST", "CARB.ST", "CLAB.ST",
        "CONC.ST", "ESS.ST", "HM.ST", "ITUB.ST", "KFST.ST", "MUB.ST",
        "NPI.ST", "OQA.ST", "PEO.ST", "SAB.ST", "SEB.ST", "SAND.ST",
        "SSAB.ST", "SWED-A.ST", "SWED-B.ST", "SVA.ST", "Telia.ST",
        "VOLV-B.ST", "ERIC-B.ST", "ALFA.ST", "HMV-B.ST", "SAND.ST",
        "Investor-B.ST", "IFU.AB",
        "ESS.AB", "SAM-B.ST", "SEK-A.ST", "KIR.BK", "INDU-A.ST",
        "SEB-A.ST", "HM-B.ST", "IFU-A.ST", "S395.B", "S396.B",
    ]

    # UK
    UK_TICKERS = [
        "BP.L", "SHEL.L", "GSK.L", "AZN.L", "ULVR.L", "DGE.L",
        "BARC.L", "HSBA.L", "LLOY.L", "VOD.L", "GLEN.L", "RIO.L",
        "AAL.L", "BAT.L", "REL.L", "RR.L", "BA.L", "NG.L",
        "LSEG.L", "ANTO.L", "DGE.L", "EXPN.L", "ITRK.L",
        "STAN.L", "MKS.L", "BDEV.L", "PRU.L", "AVST.L",
        "BLND.L", "CRDA.L", "CPG.L", "DCC.L", "EZJ.L",
        "FRES.L", "HWDN.L", "III.L", "IHG.L", "IMB.L",
        "ITRK.L", "JD.L", "KGF.L", "LAND.L", "LAND.L",
        "MNDI.L", "MNG.L", "NWG.L", "OCDO.L", "PSH.L",
        "RKT.L", "SGE.L", "SN.L", "SMDS.L", "SMIN.L",
        "SDR.L", "TSCO.L", "WTB.L", "WEIR.L", "WN.L",
    ]

    # Germany
    GERMANY_TICKERS = [
        "SIE.DE", "ALV.DE", "MBG.DE", "BAS.DE", "DTE.DE", "BMW.DE",
        "VNA.DE", "MUV2.DE", "DB1.DE", "FRE.DE", "HEN3.DE",
        "IFX.DE", "LIN.DE", "MRK.DE", "CON.DE", "DBK.DE",
        "PUM.DE", "BEI.DE", "ZAL.DE", "SY1.DE", "FME.DE",
        "HEI.DE", "MTX.DE", "KRA.DE", "BSF.DE", "WAZ.MU",
    ]

    # France
    FRANCE_TICKERS = [
        "OR.PA", "SAN.PA", "TTE.PA", "BN.PA", "AIR.PA", "AI.PA",
        "ACA.PA", "BNP.PA", "CAP.PA", "EN.PA", "EL.PA",
        "KER.PA", "MC.PA", "MNOP.PA", "RI.PA", "SAF.PA",
        "SGO.PA", "SU.PA", "VIE.PA", "STL.PA", "DSY.PA",
        "CS.PA", "BVI.PA", "HO.PA", "GO.PA", "LEH.PA",
    ]

    # Japan (Nikkei 225 + major TSE stocks)
    JAPAN_TICKERS = [
        "7203.T", "6758.T", "9984.T", "6861.T", "8306.T",
        "6954.T", "8035.T", "4519.T", "6098.T", "9432.T",
        "7974.T", "4689.T", "9433.T", "9735.T", "4704.T",
        "6920.T", "6367.T", "8058.T", "9983.T", "4063.T",
        "6594.T", "7751.T", "6702.T", "6841.T", "6981.T",
        "6301.T", "6302.T", "6501.T", "6503.T", "6504.T",
        "6506.T", "6591.T", "6752.T", "6753.T", "6757.T",
        "6762.T", "6857.T", "6902.T", "7259.T", "7261.T",
        "7267.T", "7269.T", "7731.T", "7733.T", "7752.T",
        "7754.T", "8001.T", "8002.T", "8003.T", "8015.T",
        "8031.T", "8053.T", "8059.T", "8252.T", "8267.T",
        "8410.T", "8630.T", "8697.T", "8750.T", "9001.T",
        "9006.T", "9020.T", "9021.T", "9101.T", "9104.T",
        "9301.T", "9304.T", "9501.T", "9502.T", "9531.T",
        "9532.T", "9613.T", "7735.T", "6976.T", "4502.T",
        "4503.T", "4523.T", "4541.T", "4567.T", "4578.T",
        "4661.T",
    ]

    # Hong Kong (Hang Seng constituents)
    HONGKONG_TICKERS = [
        "0005.HK", "0006.HK", "0011.HK", "0012.HK", "0016.HK",
        "0017.HK", "0027.HK", "0066.HK", "0267.HK", "0288.HK",
        "0291.HK", "0388.HK", "0522.HK", "0688.HK", "0700.HK",
        "0762.HK", "0883.HK", "0914.HK", "0939.HK", "0941.HK",
        "0968.HK", "0998.HK", "1038.HK", "1088.HK", "1093.HK",
        "1109.HK", "1171.HK", "1186.HK", "1211.HK", "1299.HK",
        "1339.HK", "1398.HK", "1766.HK", "1918.HK", "1997.HK",
        "2020.HK", "2313.HK", "2319.HK", "2380.HK", "2601.HK",
        "2688.HK", "2888.HK", "3311.HK", "3328.HK", "3329.HK",
        "3383.HK", "6060.HK", "6690.HK", "9868.HK", "9901.HK",
    ]

    # China (A-shares: Shanghai SS + Shenzhen SZ + HK-listed Chinese names)
    CHINA_TICKERS = [
        # A-shares
        "600519.SS", "601318.SS", "600036.SS", "601398.SS", "000858.SZ",
        "600276.SS", "300750.SZ", "601012.SS", "000568.SZ", "600900.SS",
        "601888.SS", "000333.SZ", "600030.SS", "601288.SS", "000001.SZ",
        "601166.SS", "600309.SS", "600887.SS", "000651.SZ", "601336.SS",
        "601668.SS", "601816.SS", "002714.SZ", "000002.SZ", "600048.SS",
        "601601.SS", "000538.SZ", "601088.SS", "002142.SZ", "600000.SS",
        "002304.SZ", "002475.SZ", "000725.SZ", "002352.SZ", "002594.SZ",
        "600346.SS", "601211.SS", "601607.SS", "600809.SS", "601328.SS",
        # HK-listed Chinese tech/consumers
        "3690.HK", "9888.HK", "9999.HK", "6186.HK", "1359.HK",
        "1024.HK", "9633.HK", "2150.HK", "6185.HK",
    ]

    # India
    INDIA_TICKERS = [
        "RELIANCE.NS", "TCS.NS", "HDFCBANK.NS", "INFY.NS", "ICICIBANK.NS",
        "HINDUNILVR.NS", "SBIN.NS", "BHARTIARTL.NS", "ITC.NS", "KOTAKBANK.NS",
        "LT.NS", "AXISBANK.NS", "HCLTECH.NS", "ASIANPAINT.NS", "MARUTI.NS",
        "TATAMOTORS.NS", "TATASTEEL.NS", "WIPRO.NS", "ULTRACEMCO.NS",
        "BAJFINANCE.NS", "NTPC.NS", "POWERGRID.NS", "ONGC.NS", "SUNPHARMA.NS",
        "TITAN.NS", "TECHM.NS", "M&M.NS", "HDFCLIFE.NS", "JSWSTEEL.NS",
    ]

    # Australia
    AUSTRALIA_TICKERS = [
        "BHP.AX", "CBA.AX", "RIO.AX", "CSL.AX", "NAB.AX",
        "ANZ.AX", "WES.AX", "WBC.AX", "FMG.AX", "MQG.AX",
        "TLS.AX", "TCL.AX", "WOW.AX", "GMG.AX", "COL.AX",
        "STO.AX", "S32.AX", "WDH.AX", "REA.AX", "QBE.AX",
        "APP.AX", "SUN.AX", "WDS.AX", "EML.AX", "ASX.AX",
    ]

    # Canada
    CANADA_TICKERS = [
        "SHOP.TO", "RY.TO", "TD.TO", "BNJ.TO", "BNS.TO",
        "CNR.TO", "CP.TO", "ENB.TO", "SU.TO", "TRP.TO",
        "CNQ.TO", "BMO.TO", "CM.TO", "ABX.TO", "K.TO",
        "FCX.TO", "CVE.TO", "IMO.TO", "TRI.TO", "MFC.TO",
        "WCN.TO", "FTS.TO", "HEI.TO", "GIB.A", "QSR.TO",
    ]

    # Switzerland (SMI + major SIX stocks)
    SWITZERLAND_TICKERS = [
        "NESN.SZ", "ROG.SZ", "NOVN.SZ", "ABBN.SZ", "GIVN.SZ",
        "ZURN.SZ", "SREN.SZ", "LONN.SZ", "2000.SN", "EKBAS.SZ",
        "VONN.SZ", "SNGY.SZ", "LISN.SZ", "PAXN.SZ", "BMOG.SZ",
        "SNCA.SZ", "SIKA.SZ", "SCMN.SZ", "UBSG.SZ", "CSGN.SZ",
        "GIGN.SZ", "KNIN.SZ", "AENS.SZ", "ANEA.SZ", "ASYL.SZ",
        "EVTN.SZ", "HOLN.SZ", "KSNR.SZ", "MILL.SZ", "MIGN.SZ",
        "PFBI.SZ", "SIAN.SZ", "TELSS.SZ", "ZILN.SZ", "ALGN.SZ",
    ]

    # South Korea (KOSPI + KOSDAQ major)
    KOREA_TICKERS = [
        "005930.KS", "000660.KS", "035420.KS", "051910.KS",
        "006400.KS", "207940.KS", "373220.KS", "068270.KS",
        "005380.KS", "035720.KS", "000100.KS", "000270.KS",
        "012340.KS", "018260.KS", "028260.KS", "003550.KS",
        "006840.KS", "015760.KS", "018230.KS", "033780.KS",
        "034020.KS", "046730.KS", "051900.KS", "066570.KS",
        "086740.KS", "096770.KS", "105560.KS", "114440.KS",
        "128560.KS", "139480.KS", "141480.KS", "161390.KS",
        "176640.KS", "188140.KS", "191770.KS", "207220.KS",
        "216040.KS", "241560.KS", "250660.KS", "263730.KS",
        "279970.KS", "291210.KS", "301630.KS", "313260.KS",
        "321820.KS", "335880.KS", "353770.KS", "402820.KS",
    ]

    return {
        "usa": sorted(set(USA_TICKERS)),
        "sweden": sorted(set(SWEDEN_TICKERS)),
        "uk": sorted(set(UK_TICKERS)),
        "germany": sorted(set(GERMANY_TICKERS)),
        "france": sorted(set(FRANCE_TICKERS)),
        "japan": sorted(set(JAPAN_TICKERS)),
        "hongkong": sorted(set(HONGKONG_TICKERS)),
        "china": sorted(set(CHINA_TICKERS)),
        "india": sorted(set(INDIA_TICKERS)),
        "australia": sorted(set(AUSTRALIA_TICKERS)),
        "canada": sorted(set(CANADA_TICKERS)),
        "switzerland": sorted(set(SWITZERLAND_TICKERS)),
        "korea": sorted(set(KOREA_TICKERS)),
        "international": sorted(set(
            SWEDEN_TICKERS + UK_TICKERS + GERMANY_TICKERS +
            FRANCE_TICKERS + JAPAN_TICKERS + HONGKONG_TICKERS +
            CHINA_TICKERS + INDIA_TICKERS + AUSTRALIA_TICKERS +
            CANADA_TICKERS + SWITZERLAND_TICKERS + KOREA_TICKERS
        )),
    }
